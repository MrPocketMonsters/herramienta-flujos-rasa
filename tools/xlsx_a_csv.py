"""Convierte hojas de un archivo XLSX a archivos CSV.

Uso:
python tools/xlsx_a_csv.py [-h] [--workbook WORKBOOK] [--output-dir OUTPUT_DIR]
                           [--sheet SHEET] [--delimiter DELIMITER] [--encoding ENCODING]

Valores por defecto:
--workbook: data/flujo_conversacional.xlsx
--output-dir: data

Nota:
La primera fila de cada hoja se considera descripcion y se omite en el CSV de salida.
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


def normalize_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def sanitize_sheet_name(sheet_name: str) -> str:
    # Evita problemas con nombres de archivos al convertir nombres de hoja.
    invalid = '<>:"/\\|?*'
    result = sheet_name
    for ch in invalid:
        result = result.replace(ch, "_")
    return result.strip()


def write_sheet_to_csv(
    workbook_path: Path,
    sheet_name: str,
    output_dir: Path,
    delimiter: str,
    encoding: str,
) -> Path:
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No se encontro la hoja '{sheet_name}' en el workbook.")

    ws = wb[sheet_name]
    output_dir.mkdir(parents=True, exist_ok=True)

    file_name = f"{sanitize_sheet_name(sheet_name)}.csv"
    out_path = output_dir / file_name

    with out_path.open("w", newline="", encoding=encoding) as f:
        writer = csv.writer(f, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx == 1:
                # La fila 1 contiene descripcion de la hoja; no se exporta al CSV.
                continue
            writer.writerow([normalize_value(cell) for cell in row])

    return out_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Convierte hojas de Excel a archivos CSV.")
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path("data/flujo_conversacional.xlsx"),
        help="Ruta al workbook XLSX",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("data"),
        help="Directorio de salida para los CSV (por defecto: data)",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Nombre de una hoja especifica a convertir. Si no se indica, convierte todas.",
    )
    parser.add_argument(
        "--delimiter",
        type=str,
        default=",",
        help="Separador CSV de salida (por defecto: coma)",
    )
    parser.add_argument(
        "--encoding",
        type=str,
        default="utf-8-sig",
        help="Encoding de salida (por defecto: utf-8-sig)",
    )
    args = parser.parse_args()

    workbook_path = args.workbook
    output_dir = args.output_dir

    if not workbook_path.exists():
        raise FileNotFoundError(f"No existe el workbook requerido: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=True)
    sheet_names = wb.sheetnames

    if args.sheet:
        if args.sheet not in sheet_names:
            raise ValueError(f"La hoja '{args.sheet}' no existe. Hojas disponibles: {sheet_names}")
        selected_sheets = [args.sheet]
    else:
        selected_sheets = sheet_names

    print(f"Workbook: {workbook_path}")
    print(f"Hojas a convertir: {len(selected_sheets)}")

    for sheet_name in selected_sheets:
        out_path = write_sheet_to_csv(
            workbook_path=workbook_path,
            sheet_name=sheet_name,
            output_dir=output_dir,
            delimiter=args.delimiter,
            encoding=args.encoding,
        )
        print(f"- OK: {sheet_name} -> {out_path}")


if __name__ == "__main__":
    main()
